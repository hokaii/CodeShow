import networkx as nx
import openpyxl

if __name__ == '__main__':
    G = nx.DiGraph()
    read = 0
    with open("web-google.txt", encoding="utf8") as f:
        for line in f:
            read += 1
            if read > 4:
                node = line.split("\t")
                node[1] = node[1].strip()
                node[0], node[1] = int(node[0]), int(node[1])
                # 添加有向边
                G.add_edge(node[0], node[1])
    pr = nx.pagerank(G, alpha=0.85, personalization=None, max_iter=100, tol=1e-06, nstart=None, weight='weight',
                     dangling=None)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0)
    column = ["Page", "PR_Value"]
    ws.cell(1, 1).value = "Node"
    ws.cell(1, 2).value = "PR"

    num = 2
    for node, pageRankValue in pr.items():
        ws.cell(num, 1).value = node
        ws.cell(num, 2).value = pageRankValue
        num += 1

    wb.save('PageRank_printout.xls')# 保存为Excel文件
    print('end!')